import re
import datetime
import os
import shutil
from openpyxl import load_workbook

TERRITORY_LIST = ['Brightwood', 'Cutlass', 'Ebonscale', 'Everfall', 'First Light', "Monarch's",
                   'Mourningdale', 'Reekwater', 'Restless', "Weaver's", 'Windsward']


def find_territory(text, file):
    territory = None
    for item in TERRITORY_LIST:
        if text.upper().find(item.upper()) != -1:    # If name appears in image use that
            territory = item
            break
        else:
            file_formatted = file.upper().replace("'", "")
            item_formatted = item.upper().replace("'", "")
            if file_formatted.find(item_formatted) != -1:    # If no name in image, search filename
                territory = item
                break
    return territory


def find_company(text):
    company = None
    t = re.findall('GOVERNED BY\n([a-zA-Z ]+)', text)
    if t: company = t[0]
    return company


def find_can_pay_in(text):
    can_pay_in = None
    # Common issues:
    #   "1" read as "I"
    t = re.findall('CAN PAY IN ([a-zA-Z0-9 ]+)', text)
    if t:
        can_pay_string = t[0].split('day')
        if len(can_pay_string) > 1:    # i.e., "Can Pay In" >= 1 day
            days = can_pay_string[0]
            hours = can_pay_string[1]
        else:
            days = '0'
            hours = can_pay_string[0]
        days = days.replace('I', '1')
        hours = hours.split('hour')[0].replace('I', '1')
        hours = re.sub('[^0-9]','', hours)
        can_pay_in = int(days) + (int(hours)+1) / 24    # Add 1 hour, as values on payroll sheet are not updated until the hour is over
    return can_pay_in


def find_due_in(text):
    due_in = None
    t = re.findall('[0-9 ]+:[ 0-9]+', text)
    if t:
        days, hours = t[0].split(':')
        due_in = int(days) + (int(hours)+1) / 24
    return due_in


def find_screenshot_date(file):
    screenshot_date = None
    p = re.findall('([2][0][0-9][0-9]-[0-1][0-9]-[0-3][0-9])', file)    # Search file name for dates formatted 20yy-mm-dd
    if p: screenshot_date = p[0]
    return screenshot_date

def find_dates(file, screenshot_date, can_pay_in, due_in):
    period_start, period_end, days_elapsed = None, None, None
    # First search the filename for period end. These dates are added manually as GCV can't tell which tab on the
    # payroll page is open. Dates follow format "Wed Oct 6" or "Wed Oct 06"
    # I've made this unnecessarily complicated for myself but wanted to match the date format used in game
    p = re.findall('([a-zA-Z][a-zA-Z][a-zA-Z]) ([a-zA-Z][a-zA-Z][a-zA-Z]) ([0-9][0-9]|[0-9])', file)
    if screenshot_date == None:
        pass
    elif p:
        # Matches the date to the closest year to the year of the screenshot, based on weekday, month, and day
        # Solution from https://stackoverflow.com/questions/36655632/calculate-year-from-weekday-month-and-day
        year = int(screenshot_date.split('-')[0])
        weekdays = {'Mon': 0, 'Tue': 1, 'Wed': 2, 'Thu': 3, 'Fri': 4, 'Sat': 5, 'Sun': 6}
        weekday = weekdays[p[0][0]]
        while year >= 2021:  # Game released in 2021.
            test_date_string = f'{year} {p[0][1]} {p[0][2]}'  # YYYY mmm dd
            test_date = datetime.datetime.strptime(test_date_string, '%Y %b %d')  # Reads YYYY mmm dd
            if test_date.weekday() == weekday:
                period_end = datetime.date.isoformat(test_date)
                days_elapsed = 7
                period_start = datetime.date.isoformat(
                    datetime.date.fromisoformat(period_end) -
                    datetime.timedelta(days_elapsed)
                )
                break
            else:
                year -= 1
    # If no period stated in filename, calculate the days elapsed/period start based on the due_in/can_pay_in parameters
    else:
        # days elapsed
        if due_in != None:
            days_elapsed = 7 - due_in
        elif can_pay_in != None:
            if can_pay_in <= 2:
                days_elapsed = 2 - can_pay_in
            else:
                days_elapsed = 2 + 7 - can_pay_in
        # Period start
        if days_elapsed != None:
            period_start = datetime.date.isoformat(
                datetime.date.fromisoformat(screenshot_date) -
                datetime.timedelta(days_elapsed)
            )

        # period end
        if period_start != None:
            period_end = datetime.date.isoformat(
                datetime.date.fromisoformat(period_start) +
                datetime.timedelta(7)
            )
    return period_start, period_end, days_elapsed


def find_tax_rates(text):
    property_tax_rate = None
    trading_tax_rate = None
    crafting_tax_rate = None
    refining_tax_rate = None

    # Property tax %; Trading tax %
    t = re.findall('[0-9]+.[0-9][0-9]%', text)
    if t:
        property_tax_rate = t[0].replace('%', '')
        trading_tax_rate = t[1].replace('%', '')

    # Crafting tax rate, refining tax rate
    # Common issues:
    #    "x" read as "kha", which looks exactly like x
    t = re.findall('[x|х]([ 0-9]+.[0-9][0-9])', text)
    if t:
        crafting_tax_rate = float(t[0])
        refining_tax_rate = float(t[1])

    return property_tax_rate, trading_tax_rate, crafting_tax_rate, refining_tax_rate


def check_extra_nine(t):
    """Common issue with Google Cloud vision is that it reads the coin symbol as a 9.
    This function checks for an extra 9 at the end of the numbers (which should end with a decimal and 2 numbers """
    check = re.findall('.[0-9][0-9][9]', t[0][-4:])
    if check != []: extra_nine = True
    else: extra_nine = False
    return extra_nine

def find_tax_income(text):
    property_income = None
    trading_income = None
    crafting_income = None
    refining_income = None
    # Property tax income
    # Common issues:
    #   "," read as "."
    #   gold coin symbol at end read as "9", and not separated from the numbers by a space.
    t = re.findall('Property Tax\n([0-9.,]+)', text)
    # "," is read as "." sometimes and it messes things up. So we remove all symbols and re-introduce them by dividing by 100
    if t:
        if check_extra_nine(t): t[0] = t[0][:-1]    # remove last digit if there is an extra 9.
        property_income = float(re.sub('[^0-9]', '', t[0])) / 100

    # Trading tax income
    t = re.findall('Trading Tax\n([0-9.,]+)', text)
    if t:
        if check_extra_nine(t): t[0] = t[0][:-1]
        trading_income = float(re.sub('[^0-9]', '', t[0])) / 100

    # Crafting tax income
    t = re.findall('Crafting Fee\n([0-9.,]+)', text)
    if t:
        if check_extra_nine(t): t[0] = t[0][:-1]
        crafting_income = float(re.sub('[^0-9]', '', t[0])) / 100

    # Refining tax income
    # Common issues:
    #   "O" read instead of 0
    #   no space between numbers and words
    t = re.findall('Refining Fee\n([0-9.,]+)', text)
    if t:
        if check_extra_nine(t): t[0] = t[0][:-1]
        refining_income = float(re.sub('[^0-9]', '', t[0])) / 100

    return property_income, trading_income, crafting_income, refining_income


def find_tax_volume(text):
    property_volume = None
    trading_volume = None
    crafting_volume = None
    refining_volume = None
    # Property volume
    t = re.findall('([0-9O ]+)housing units', text)    # Include "O" as sometimes 0 is read as O
    if t: property_volume = int(t[0].replace('O', '0'))

    # Trading volume
    t = re.findall('([0-9O ]+)transactions', text)
    if t: trading_volume = int(t[0].replace('O', '0'))

    # Crafting volume
    t = re.findall('([0-9O ]+)items crafted', text)
    if t: crafting_volume = int(t[0].replace('O', '0'))

    # Refining volume
    t = re.findall('([0-9O ]+)resources refined', text)
    if t: refining_volume = int(t[0].replace('O', '0'))

    return property_volume, trading_volume, crafting_volume, refining_volume


def extract_variables(text, file):
    """
    Only keeps track of the first instance of each variable. Eg., having information for multiple cities / time periods
     on a single image/text will only return the first city/period. Each payroll page should not be saved in the same
     image as other payroll pages for that reason.
    If a value is missing it will be recorded as None (i.e., empty cell in Excel).
    """
    territory = find_territory(text, file)
    company = find_company(text)
    can_pay_in = find_can_pay_in(text)
    due_in = find_due_in(text)
    screenshot_date = find_screenshot_date(file)
    period_start, period_end, days_elapsed = find_dates(file, screenshot_date, can_pay_in, due_in)
    property_tax_rate, trading_tax_rate, crafting_tax_rate, refining_tax_rate = find_tax_rates(text)
    property_income, trading_income, crafting_income, refining_income = find_tax_income(text)
    property_volume, trading_volume, crafting_volume, refining_volume = find_tax_volume(text)

    row = {
        'Territory' : territory,
        'Faction' : None,    # Shown by colour and can't be read by GCV
        'Company' : company,
        'Period Start': period_start,
        'Period End': period_end,
        'Screenshot Date': screenshot_date,
        'Can Pay In' : can_pay_in,
        'Due In' : due_in,
        'Days Elapsed' : days_elapsed,
        'Property %' : property_tax_rate,
        'Trading %': trading_tax_rate,
        'Crafting %': crafting_tax_rate,
        'Refining %': refining_tax_rate,
        'Property #' : property_volume,
        'Property $' : property_income,
        'Trading #': trading_volume,
        'Trading $': trading_income,
        'Crafting #': crafting_volume,
        'Crafting $': crafting_income,
        'Refining #': refining_volume,
        'Refining $': refining_income,
        'Total $' : None,    # Calculate in spreadsheet
        '$ per day' : None    # Calculate in spreadsheet
    }
    return row


def open_text(path):
    with open(path, 'r', encoding='utf-8') as file:
        text = file.read()
    return text


def texts_to_df(df):
    files = os.listdir('texts/unread')
    if len(files) > 0:
        for file in files:
            text = open_text(f'texts/unread/{file}')
            row = extract_variables(text, file)
            df = df.append(row, ignore_index=True)
            shutil.move('texts/unread/' + file, 'texts/read/' + file)
            print('Finished reading ' + file)
    return df


def texts_to_excel(source_file):
    files = os.listdir('texts/unread')
    if len(files) > 0:
        for file in files:

            # Open text and extract variables
            text = open_text(f'texts/unread/{file}')
            row = extract_variables(text, file)

            # Add row to excel sheet. Solution copied from:
            # https://stackoverflow.com/questions/65894853/how-to-extend-the-range-of-an-excel-table-in-openpyxl
            book = load_workbook(source_file)
            tb = book.active.tables['Table1']
            curr_ref = tb.ref
            values = list(row.values())
            book.active.append(values)
            tb.ref = re.sub(r"\d+$", str(book.active.max_row), curr_ref)  # Extend table range. Copied from SO
            book.save(source_file)

            # move read text files to read folder
            shutil.move('texts/unread/' + file, 'texts/read/' + file)
            print('Finished reading ' + file)
    else:
        print('No files in directory.')


if __name__ == '__main__':
    os.chdir('../..')
    source_file = 'Server Economy.xlsx'
    texts_to_excel(source_file)



# Test a single file:
# if False:
#     file = "2021-10-11 Monarch's Mon Oct 11.txt"
    # text = open_text(f"texts/unread/{file}")
    # row = extract_variables(text, file)